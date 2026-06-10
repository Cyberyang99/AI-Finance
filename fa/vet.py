"""逻辑校验器 (vet) — 用已有 CoT + note 审视一只股票 / 一个投资想法 / 一份标的清单.

定位：合成层。区别于「建仓→回顾→进化」闭环，这里不做业绩验证，只做知识检索 + 逻辑校验。

三种入口：
  - vet_stock(ticker, idea)：单股全量校验。拉基本面快照 → LLM 召回 CoT+note →
    合成校验笔记（打分/补漏/反逻辑/同业共性）→ markdown 落盘
  - vet_idea(idea)：观点-only（无标的）。按观点主题召回 → 校验+标的映射 → markdown 落盘
  - vet_batch(items)：标的清单轻量扫描。目录一次加载，每股 1 次 LLM 调用
    （召回+短结论合一）→ 汇总 Excel 落盘

idea 均可为内联文本或 word/pdf/pptx/txt/md 文件路径。
输出落盘到 output_dir（默认 ~/Desktop），**不自动入 note 库**。
"""

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from .config import load_config, make_anthropic_client
from .cot.loader import load_cots
from .ingest.base import SUPPORTED_EXT
from .ingest.user_note import load_user_notes
from .tools.data import fetch_fundamentals


def _model() -> str:
    """合成用模型：复用 [cot] extract_model（当前 flash），否则跟 [agent] model。"""
    cfg = load_config()
    return cfg.get("cot", {}).get("extract_model") or cfg.get("agent", {}).get("model", "deepseek-v4-flash")


def _output_dir() -> Path:
    cfg = load_config().get("paths", {})
    return Path(cfg.get("output_dir", "~/Desktop")).expanduser()


def _parse_json(text: str) -> Optional[dict]:
    if not text:
        return None
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass
    s, e = text.find("{"), text.rfind("}")
    if s >= 0 and e > s:
        try:
            return json.loads(text[s:e + 1])
        except Exception:
            return None
    return None


def load_idea(idea: str) -> str:
    """idea 可以是内联文本，或一个文件路径（docx/pdf/pptx/xlsx 走抽文器；txt/md 直接读）。"""
    if not idea or not idea.strip():
        return ""
    from .ingest.base import clean_user_path
    p = Path(clean_user_path(idea)).expanduser()
    if p.exists() and p.is_file():
        ext = p.suffix.lower()
        if ext in SUPPORTED_EXT:
            from .ingest import ingest_file
            try:
                return ingest_file(p)["text"]
            except Exception as ex:
                print(f"  [vet] 想法文件抽文失败({ex})，当作纯文本忽略")
                return ""
        try:
            return p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return ""
    return idea  # 内联文本


def _fundamentals_snapshot(d: dict) -> str:
    lines = [f"代码: {d.get('ticker','?')}  名称: {d.get('name','?')}  行业: {d.get('sector') or '未知'}"]
    parts = []
    # 市值显式折成 亿+币种：裸的原始大数（如 506477750240）LLM 容易读错量级
    mc_yi = d.get("market_cap_yi") or (round(d["market_cap"] / 1e8, 1) if d.get("market_cap") else None)
    if mc_yi:
        parts.append(f"市值={mc_yi}亿{d.get('currency', '')}")
    metrics = [
        ("营收3yCAGR%", "revenue_cagr_3y"), ("毛利率%", "gross_margin"),
        ("净利率%", "net_margin"), ("ROE%", "roe"), ("负债率%", "debt_ratio"),
        ("PE", "pe"), ("股息率%", "div_yield"),
    ]
    parts += [f"{label}={round(d[key], 2) if isinstance(d[key], float) else d[key]}"
              for label, key in metrics if d.get(key) is not None]
    if parts:
        lines.append("  ".join(parts))
    return "\n".join(lines)


# ── 紧凑目录（召回/批量共用；C#/N# 编号即列表下标） ──

def _cot_catalog(cots: list) -> str:
    return "\n".join(
        f"C{i}: (信号{c.get('signal','?')}|{c.get('_sector','')}|{'/'.join(c.get('_tags', []))}) {c['trigger']}"
        for i, c in enumerate(cots)
    )


def _note_catalog(notes: list) -> str:
    return "\n".join(
        f"N{i}: [{n['ticker']} | {n.get('sector','?')} | {'/'.join(n.get('tags', [])) or '无标签'}] "
        f"{re.sub(chr(92)+'s+', ' ', n['content'])[:130]}"
        for i, n in enumerate(notes)
    )


# ── 召回 ──

RECALL_SYSTEM = """你是投资知识库的检索助手。给你一只目标股票（含基本面）、用户的投资想法（可能为空），
以及全库 CoT（行业可复用逻辑）和【同业候选 note】（其它个股的研究笔记，已剔除目标公司自身）的紧凑目录。
你的任务：按语义相关性挑出最有用的 CoT，以及**最可比的同业 note**（同行业 / 同主题标签优先，用于横向对比）。
重点：note 要挑「能跟目标公司做对比、提炼共性的同类公司」，不是只看名字像。宁缺毋滥，但别漏掉强可比对象。
同一主题若存在结论相反的 CoT，**两边都要挑进来**——冲突链是校验的关键弹药，单边召回会让校验失真。"""

RECALL_TEMPLATE = """## 目标股票
{snapshot}

## 用户想法（可能为空）
{idea}

## 候选 CoT 目录（编号 C#）
{cot_catalog}

## 同业候选 note 目录（编号 N#，已剔除目标公司自己）
{note_catalog}

## 输出（严格 JSON，无 markdown 包裹）
{{
  "cot_idx": [相关 CoT 的编号数字, 最多 {cot_limit} 个, 按相关性排序],
  "note_idx": [可比同业 note 的编号数字, 最多 {note_limit} 个, 同行业/同标签优先],
  "reasoning": "<一句话说明为什么挑这些、可比在哪>"
}}
只输出 JSON。"""


def _recall(snapshot: str, idea: str, cots: list, notes: list,
            cot_limit: int, note_limit: int, model: str, client) -> tuple[list, list, str]:
    cot_cat = _cot_catalog(cots)
    note_cat = _note_catalog(notes)
    msg = RECALL_TEMPLATE.format(
        snapshot=snapshot, idea=idea or "（无，做产业逻辑命中扫描）",
        cot_catalog=cot_cat or "（空）", note_catalog=note_cat or "（空）",
        cot_limit=cot_limit, note_limit=note_limit,
    )
    try:
        # max_tokens 含 thinking 预算：deepseek-v4 思考模式下 thinking block 先吃配额，
        # 给小了 text 为空（stop_reason=max_tokens）、JSON 解析必败
        resp = client.messages.create(model=model, max_tokens=6000, system=RECALL_SYSTEM,
                                      messages=[{"role": "user", "content": msg}])
        text = "".join(b.text for b in resp.content if b.type == "text")
        stop = getattr(resp, "stop_reason", "?")
    except Exception as e:
        print(f"  [vet] ⚠ 召回 LLM 失败({e})，回退：取信号最高的 CoT（结果可能偏离主题）")
        top = sorted(range(len(cots)), key=lambda i: -int(cots[i].get("signal", 5)))[:cot_limit]
        return top, list(range(min(len(notes), note_limit))), "（召回失败，按信号兜底）"

    parsed = _parse_json(text) or {}
    cot_idx = [i for i in parsed.get("cot_idx", []) if isinstance(i, int) and 0 <= i < len(cots)]
    note_idx = [i for i in parsed.get("note_idx", []) if isinstance(i, int) and 0 <= i < len(notes)]
    if not cot_idx:  # 兜底必须出声：静默降级会让合成拿一堆高信号但跑题的链
        print(f"  [vet] ⚠ 召回解析失败或为空（stop_reason={stop}, text {len(text)} 字），"
              f"回退信号兜底——结果可能偏离主题，建议重跑")
        cot_idx = sorted(range(len(cots)), key=lambda i: -int(cots[i].get("signal", 5)))[:cot_limit]
        return cot_idx, note_idx[:note_limit], "（召回解析失败，按信号兜底）"
    return cot_idx[:cot_limit], note_idx[:note_limit], parsed.get("reasoning", "")


# ── 合成 ──

SYNTH_SYSTEM = """你是中性、客观、就事论事的投资分析师。立场不预设多空——只对照已沉淀的行业逻辑(CoT)和同业研究(note)，
判断每条逻辑是否站得住，再补充你自己从知识库里读出的新角度。

铁律：
1. **中性客观**：不附和、不唱衰。证据支持就说成立，证据不足就说存疑，没证据就说「库内无支撑」。
2. **逆 CoT 即风险**：如果用户的某条逻辑与某条 CoT 的结论相反/矛盾，**这本身就是一个潜在风险点**，必须在风险一节明确点出「你的判断 vs CoT 学到的，谁可能错」。
3. **CoT 之间冲突必须配对呈现**：召回的 CoT 里若存在结论相反的两派（如「2B 不可持续」vs「B 端按需付费加速」），必须并列引用、写明各自前提和分歧点，**禁止单边引用抬高某一方置信度**。
4. **同源自证要降权**：每条 CoT 标了「来源」。若来源明显是目标公司自己的材料（公司名/高管访谈/官方报告），命中只说明「公司叙事自洽」，**不构成独立佐证**——引用时标「⚠同源」，逐条校验的可信度不得主要靠同源链支撑。
5. **引用要实**：每个判断都落到具体 CoT 编号(C#)+信号强度，或具体同业 note(写公司代码)，禁止泛泛而谈。CoT 自带「证伪条件」「原文依据」，是写风险和反证的弹药。
6. **先校验、后创造**：先逐条审用户的假设；然后**主动提出用户没说、但 CoT/note 支持的新观点**。
7. **横向对比**：用同业 note 提炼可比公司的共性与差异，不要只盯目标公司一家。
8. 没有用户想法时，转为「该股可能命中哪些产业逻辑 + 同业共性 + 主要风险」的扫描。"""

SYNTH_TEMPLATE = """## 目标股票基本面
{snapshot}

## 用户的投资想法（可能为空）
{idea}

## 用户对该公司已有的 note（自己的既有观点，按日期新→旧；勿当同业对比）
{own_notes}
（注：最新一条 = 用户「当前观点」；若多条且立场有变化，在结论里点出「观点演变」。）

## 召回的相关 CoT（行业可复用逻辑，含信号/证伪/原文依据）
{cots}

## 召回的【同业】note（其它可比公司，用于横向对比）
{notes}

---

输出一份 **markdown 校验笔记**，风格简洁、每条逻辑独立成段、善用 ✅/🟡/❌/⚠️ 与表格。结构如下（不要加代码块包裹）：

# 🔍 {ticker} 逻辑校验

## 一句话结论
结论 + 「逻辑整体可信度 X/10」（无用户想法则给「产业逻辑契合度」）。

## 逐条校验（你的逻辑）
把用户想法拆成「逻辑一/二/三…」，每条一段：
- 标题行用 ✅成立 / 🟡部分成立 / ❌存疑 / ⚠️库内无支撑 标注
- 叙述为什么，**引用具体 C# + 信号**或同业 note
- 每条末尾一句「结论：…」
（若无用户想法，本节改为「产业逻辑命中扫描」：该股可能命中哪些 CoT、强弱如何）

## 新增观点（你没提，但知识库支持）
基于 CoT/note 主动补 1-3 条用户没覆盖的角度或驱动，引用来源。

## 横向同业对比
表格：可比公司 | 共同逻辑/护城河 | 与目标的差异 | 来源note。提炼同类公司共性。

## 反逻辑与风险点
- 最强的看空/证伪理由（2-4 条），引用 CoT 证伪条件
- **⚠ 与知识库相悖之处**：用户逻辑里与某条 CoT 结论相反的地方（若有），点明分歧
- 崩溃的可观测信号：什么数据/事件出现就说明逻辑错了

## 综合
一个小表格（逻辑链 | 置信度 | 关键抓手）+ 一句话总结核心观测点。
{todo_section}
要求：引用具体到 C#/note 代码；中性克制不堆砌。不要自行添加上述结构之外的章节。"""

# 独立跑 vet 时的尾巴：提示哪些分析还没做（fa report 里有统一后段，不生成这节）
TODO_SECTION = """
## 待补充（后续接入）
- [ ] 基本面深析　- [ ] 估值（DCF/PE）　- [ ] 预期分析（可验证可进化）
"""


def _fmt_cots(cots: list, idx: list) -> str:
    out = []
    for i in idx:
        c = cots[i]
        ev = (c.get("evidence") or "").strip()
        src = (c.get("_source") or "").strip()
        out.append(
            f"【C{i}】信号{c.get('signal','?')} "
            f"(传导{c.get('transmission','?')}/证伪{c.get('falsifiability','?')}/历史{c.get('history','?')}/时效{c.get('recency','?')}) "
            f"[{c.get('_sector','')}|{'/'.join(c.get('_tags', []))}]\n"
            f"  逻辑: {c['trigger']}\n"
            f"  传导链: {c['COT']}\n"
            + (f"  来源: {src}\n" if src else "")
            + (f"  原文依据: 「{ev}」\n" if ev else "")
        )
    return "\n".join(out) or "（无相关 CoT）"


def _fmt_notes(notes: list, idx: list, max_chars: int = 1200) -> str:
    out = []
    for i in idx:
        n = notes[i]
        body = re.sub(r"\n{3,}", "\n\n", n["content"]).strip()[:max_chars]
        out.append(f"【note {n['ticker']} {n['created_at']}】\n{body}\n")
    return "\n".join(out) or "（无相关同业 note）"


def vet_stock(ticker: str, idea: str = "", cot_limit: int = 15, note_limit: int = 5,
              save: bool = True, with_todo: bool = True) -> dict:
    """主入口。返回 {ticker, markdown, path, recall_reasoning, n_cots, n_notes}。

    with_todo=False 时不生成「待补充」尾节（fa report 调用时用——报告自带统一估值后段，
    让 LLM 压根不写比事后正则剥除可靠：LLM 会给标题加 emoji 等装饰，正则防不住）。
    """
    ticker = ticker.upper().strip()
    model = _model()
    client = make_anthropic_client()

    print(f"[VET] {ticker} — 拉基本面...")
    data = fetch_fundamentals(ticker) or {"ticker": ticker}
    snapshot = _fundamentals_snapshot(data)

    idea_text = load_idea(idea)
    if idea_text:
        print(f"  ✓ 想法输入 {len(idea_text)} 字")

    cots = load_cots()
    all_notes = load_user_notes()
    own_notes = load_user_notes(ticker=ticker)          # 目标公司自己的 note（背景）
    own_paths = {n["path"] for n in own_notes}
    peer_notes = [n for n in all_notes if n["path"] not in own_paths]  # 同业候选（横向对比）
    print(f"  知识库: {len(cots)} 条 CoT / 自有 note {len(own_notes)} 份 / 同业候选 {len(peer_notes)} 份 → 自然语言召回...")

    cot_idx, peer_idx, why = _recall(snapshot, idea_text, cots, peer_notes,
                                     cot_limit, note_limit, model, client)
    print(f"  ✓ 召回 {len(cot_idx)} 条 CoT / {len(peer_idx)} 份同业 note  {('— ' + why) if why else ''}")

    cots_block = _fmt_cots(cots, cot_idx)
    notes_block = _fmt_notes(peer_notes, peer_idx)
    own_block = _fmt_notes(own_notes, range(len(own_notes))) if own_notes else "（无）"
    msg = SYNTH_TEMPLATE.format(
        ticker=ticker, snapshot=snapshot,
        idea=idea_text or "（无，做产业逻辑命中扫描）",
        own_notes=own_block,
        cots=cots_block,
        notes=notes_block,
        todo_section=TODO_SECTION if with_todo else "",
    )
    print(f"  合成校验笔记中...")
    from .framework import inject_review_rules
    try:
        resp = client.messages.create(model=model, max_tokens=6000,
                                      system=inject_review_rules(SYNTH_SYSTEM),
                                      messages=[{"role": "user", "content": msg}])
        md = "".join(b.text for b in resp.content if b.type == "text").strip()
    except Exception as e:
        return {"ticker": ticker, "error": f"合成失败: {e}", "markdown": "", "path": None}

    # 去掉可能的代码块包裹
    md = re.sub(r"^```(?:markdown)?\s*|\s*```$", "", md).strip()

    path = None
    if save and md:
        out_dir = _output_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        safe = re.sub(r"[\\/:*?\"<>|]", "_", ticker)
        path = out_dir / f"vet_{safe}_{date.today().isoformat()}.md"
        path.write_text(md, encoding="utf-8")

    return {
        "ticker": ticker, "markdown": md, "path": str(path) if path else None,
        "recall_reasoning": why, "n_cots": len(cot_idx), "n_notes": len(peer_idx),
        # 下游复用的上下文（fa report 的框架分析不重跑召回）
        "_snapshot": snapshot, "_name": data.get("name") or "",
        "_cots_block": cots_block, "_notes_block": notes_block,
        "_own_notes_block": own_block, "_idea_text": idea_text,
        "_fund": {k: data.get(k) for k in ("pe", "pb", "market_cap", "currency", "div_yield")},
    }


# ── 观点-only 模式（无标的，按主题召回） ──

IDEA_RECALL_SYSTEM = """你是投资知识库的检索助手。给你用户的一个投资观点/逻辑（无特定标的），
以及全库 CoT（行业可复用逻辑）和个股 note 的紧凑目录。
任务：围绕观点的**主题**挑出最相关的 CoT（支持或反对都算相关），以及逻辑上挂得上钩的个股 note（用于把观点映射到具体标的）。
宁缺毋滥，但结论相反的 CoT 一定要挑进来——那是校验观点的关键弹药。"""

IDEA_RECALL_TEMPLATE = """## 用户观点
{idea}

## 候选 CoT 目录（编号 C#）
{cot_catalog}

## 个股 note 目录（编号 N#）
{note_catalog}

## 输出（严格 JSON，无 markdown 包裹）
{{
  "cot_idx": [相关 CoT 的编号数字, 最多 {cot_limit} 个, 按相关性排序],
  "note_idx": [相关 note 的编号数字, 最多 {note_limit} 个],
  "reasoning": "<一句话：观点主题是什么、为什么挑这些>"
}}
只输出 JSON。"""

IDEA_SYNTH_SYSTEM = """你是中性、客观、就事论事的投资分析师。用户给出一个投资观点（无特定标的），
你对照已沉淀的行业逻辑(CoT)和个股研究(note)逐条校验。

铁律：
1. **中性客观**：证据支持就说成立，证据不足就说存疑，没证据就说「库内无支撑」。
2. **逆 CoT 即风险**：观点与某条 CoT 结论相反，本身就是风险点，必须点明分歧。
3. **引用要实**：判断落到具体 C#（含信号强度）或 note（写公司代码），禁止泛泛而谈。
4. **先校验、后创造**：先逐条审观点；再主动补用户没说、但库内支持的相关逻辑。
5. **落到标的**：观点能映射到哪些已研究的公司，明确写出来。"""

IDEA_SYNTH_TEMPLATE = """## 用户观点
{idea}

## 召回的相关 CoT（行业可复用逻辑，含信号/证伪/原文依据）
{cots}

## 召回的相关 note（个股研究，用于把观点落到标的）
{notes}

---

输出一份 **markdown 校验笔记**（不要代码块包裹），结构：

# 🔍 观点校验

## 一句话结论
结论 + 「观点可信度 X/10」。

## 逐条校验
把观点拆成「逻辑一/二/三…」，每条一段：
- 标题行用 ✅成立 / 🟡部分成立 / ❌存疑 / ⚠️库内无支撑 标注
- 叙述为什么，**引用具体 C# + 信号**或 note
- 每条末尾一句「结论：…」

## 新增观点（你没提，但知识库支持）
补 1-3 条用户没覆盖的相关角度或驱动，引用来源。

## 标的映射
表格：相关标的 | 挂钩的逻辑 | 方向(受益/受损) | 来源。库内 note 没覆盖但 CoT 点名的公司也可列。

## 反逻辑与风险点
- 最强的反对理由（2-4 条），引用 CoT 证伪条件
- **⚠ 与知识库相悖之处**：观点与某条 CoT 结论相反的地方（若有），点明分歧
- 观点证伪的可观测信号：什么数据/事件出现就说明观点错了

## 综合
一个小表格（逻辑链 | 置信度 | 关键抓手）+ 一句话总结核心观测点。

要求：引用具体到 C#/note 代码；中性克制不堆砌。"""


def vet_idea(idea: str, cot_limit: int = 15, note_limit: int = 5,
             tag: str = "", save: bool = True) -> dict:
    """观点-only 入口：无标的，按观点主题召回。返回 {markdown, path, recall_reasoning, n_cots, n_notes}。"""
    model = _model()
    client = make_anthropic_client()

    idea_text = load_idea(idea)
    if not idea_text.strip():
        return {"error": "观点为空（-i 给内联文本或文件路径）", "markdown": "", "path": None}
    print(f"[VET] 观点校验 — 输入 {len(idea_text)} 字")

    canon = None
    if tag:
        from .sectors import resolve_theme_tag
        canon, cands = resolve_theme_tag(tag)
        if not canon:
            hint = f"，相近候选: {', '.join(cands)}" if cands else ""
            return {"error": f"主题 '{tag}' 不在闭合词表里{hint}（fa sectors 查看全部）",
                    "markdown": "", "path": None}

    cots = load_cots(tag=canon) if canon else load_cots()
    notes = load_user_notes()
    scope = f"（主题过滤: {canon}）" if canon else ""
    print(f"  知识库: {len(cots)} 条 CoT{scope} / note {len(notes)} 份 → 自然语言召回...")

    msg = IDEA_RECALL_TEMPLATE.format(
        idea=idea_text[:3000], cot_catalog=_cot_catalog(cots) or "（空）",
        note_catalog=_note_catalog(notes) or "（空）",
        cot_limit=cot_limit, note_limit=note_limit,
    )
    parsed = {}
    try:
        resp = client.messages.create(model=model, max_tokens=6000, system=IDEA_RECALL_SYSTEM,
                                      messages=[{"role": "user", "content": msg}])
        text = "".join(b.text for b in resp.content if b.type == "text")
        parsed = _parse_json(text) or {}
    except Exception as e:
        print(f"  [vet] 召回 LLM 失败({e})，回退：取信号最高的 CoT")
    cot_idx = [i for i in parsed.get("cot_idx", []) if isinstance(i, int) and 0 <= i < len(cots)][:cot_limit]
    note_idx = [i for i in parsed.get("note_idx", []) if isinstance(i, int) and 0 <= i < len(notes)][:note_limit]
    if not cot_idx:  # 兜底必须出声
        print("  [vet] ⚠ 召回解析失败或为空，回退信号兜底——结果可能偏离主题，建议重跑")
        cot_idx = sorted(range(len(cots)), key=lambda i: -int(cots[i].get("signal", 5)))[:cot_limit]
    why = parsed.get("reasoning", "")
    print(f"  ✓ 召回 {len(cot_idx)} 条 CoT / {len(note_idx)} 份 note  {('— ' + why) if why else ''}")

    msg = IDEA_SYNTH_TEMPLATE.format(idea=idea_text, cots=_fmt_cots(cots, cot_idx),
                                     notes=_fmt_notes(notes, note_idx))
    print("  合成校验笔记中...")
    from .framework import inject_review_rules
    try:
        resp = client.messages.create(model=model, max_tokens=6000,
                                      system=inject_review_rules(IDEA_SYNTH_SYSTEM),
                                      messages=[{"role": "user", "content": msg}])
        md = "".join(b.text for b in resp.content if b.type == "text").strip()
    except Exception as e:
        return {"error": f"合成失败: {e}", "markdown": "", "path": None}
    md = re.sub(r"^```(?:markdown)?\s*|\s*```$", "", md).strip()

    path = None
    if save and md:
        out_dir = _output_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        slug = re.sub(r"[\\/:*?\"<>|\s]+", "-", idea_text[:16]).strip("-") or "idea"
        path = out_dir / f"vet_idea_{slug}_{date.today().isoformat()}.md"
        path.write_text(md, encoding="utf-8")

    return {"markdown": md, "path": str(path) if path else None,
            "recall_reasoning": why, "n_cots": len(cot_idx), "n_notes": len(note_idx)}


# ── 批量轻量扫描（清单 → Excel） ──

BATCH_SYSTEM = """你是中性、客观的投资分析师，做批量「产业逻辑命中扫描」（轻量版）。
给你全库 CoT 目录（行业可复用逻辑）与个股 note 目录，再给一只目标股票（基本面快照 + 可选用户观点）。
任务：挑出与该股最相关的 CoT、给一句话结论与产业逻辑契合度分（0-10）、点出主要风险/反逻辑、列最可比同业。
要求：中性克制；引用 CoT 用编号；只输出 JSON。"""

# 注意：目录块必须放最前且逐股不变（DeepSeek 按前缀自动缓存，批量时只为快照部分付全价）
BATCH_TEMPLATE = """## 候选 CoT 目录（编号 C#）
{cot_catalog}

## 个股 note 目录（编号 N#）
{note_catalog}

## 目标股票
{snapshot}

## 用户观点（可能为空）
{idea}

## 输出（严格 JSON，无 markdown 包裹）
{{
  "fit_score": <0-10 数字，产业逻辑契合度>,
  "verdict": "<一句话结论，覆盖方向与最强驱动>",
  "hits": [{{"idx": <CoT 编号数字>, "comment": "<为何命中/对该股意味着什么，一句话>"}}, 最多 {top} 个, 按相关性排序],
  "risks": "<主要风险与反逻辑，1-2 句，有证伪条件就引用>",
  "peers": ["<最可比同业的 ticker，来自 note 目录>", 最多 3 个]
}}
只输出 JSON。"""

_TICKER_HEADERS = {"ticker", "代码", "股票代码", "标的", "symbol", "股票"}
_IDEA_HEADERS = {"观点", "想法", "逻辑", "idea", "thesis", "论点"}


def parse_batch_input(arg: str) -> tuple[list[dict], list[str]]:
    """解析 --batch 参数：xlsx/csv/txt 清单文件，或逗号/空白分隔的内联清单。

    返回 (items, warnings)；items 元素 {"ticker", "idea"}，ticker 已标准化、按序去重。
    列识别：表头含 ticker/代码/标的/symbol → ticker 列；观点/想法/逻辑/idea → idea 列。
    无表头的表格只读第 1 列（观点列必须靠表头声明）；txt 每行 = ticker [分隔符 观点]。
    无交易所后缀的输入走模糊解析（联网，慢），解析失败跳过并提示。
    """
    from .chat.resolver import _normalize_ticker
    from .ingest.base import clean_user_path

    raw: list[tuple[str, str]] = []
    p = Path(clean_user_path(arg)).expanduser()
    if p.exists() and p.is_file():
        ext = p.suffix.lower()
        if ext in (".xlsx", ".xlsm", ".csv"):
            if ext == ".csv":
                import csv
                with open(p, newline="", encoding="utf-8-sig") as f:
                    grid = [[c.strip() for c in r] for r in csv.reader(f)]
            else:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), data_only=True, read_only=True)
                ws = wb[wb.sheetnames[0]]
                grid = [[("" if v is None else str(v).strip()) for v in r]
                        for r in ws.iter_rows(values_only=True)]
                wb.close()
            grid = [g for g in grid if any(g)]
            if not grid:
                return [], [f"{p.name} 是空表"]
            head = [c.lower() for c in grid[0]]
            t_col = next((i for i, h in enumerate(head) if h in _TICKER_HEADERS), None)
            i_col = next((i for i, h in enumerate(head) if h in _IDEA_HEADERS), None)
            body = grid[1:] if (t_col is not None or i_col is not None) else grid
            t_col = t_col if t_col is not None else 0
            for g in body:
                tk = g[t_col] if t_col < len(g) else ""
                idea = g[i_col] if (i_col is not None and i_col < len(g)) else ""
                if tk:
                    raw.append((tk, idea))
        else:  # txt/md：每行一个标的，后可跟观点（空白/逗号分隔）
            for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = re.split(r"[\s,，;；]+", line, maxsplit=1)
                raw.append((parts[0], parts[1].strip() if len(parts) > 1 else ""))
    else:
        raw = [(t, "") for t in re.split(r"[,，;；\s]+", arg) if t.strip()]

    items: list[dict] = []
    warns: list[str] = []
    seen: set[str] = set()
    for tk, idea in raw:
        if "." in tk:
            norm = _normalize_ticker(tk)
        else:
            from .chat.resolver import resolve_one
            hit = resolve_one(tk)
            if not hit:
                warns.append(f"'{tk}' 无法解析为标的，已跳过")
                continue
            norm = hit["ticker"]
            warns.append(f"'{tk}' → {norm} ({hit.get('name', '')})")
        if norm in seen:
            continue
        seen.add(norm)
        items.append({"ticker": norm, "idea": idea})
    return items, warns


def vet_batch(items: list[dict], top: int = 5, save: bool = True) -> dict:
    """批量轻量扫描：目录一次加载，每标的 1 次 LLM 调用（召回+短结论合一），汇总写 Excel。

    items: [{"ticker", "idea"}]。单股失败不中断，错误记进该行「备注」列。
    返回 {path, results}。
    """
    model = _model()
    client = make_anthropic_client()

    cots = load_cots()
    notes = load_user_notes()
    cot_cat = _cot_catalog(cots) or "（空）"
    note_cat = _note_catalog(notes) or "（空）"
    # 循环外算一次：system 逐股不变，DeepSeek 前缀缓存不被打破
    from .framework import inject_review_rules
    batch_system = inject_review_rules(BATCH_SYSTEM)
    print(f"[VET] 批量扫描 {len(items)} 只 — 知识库: {len(cots)} 条 CoT / {len(notes)} 份 note（目录一次加载）")

    rows: list[dict] = []
    hit_rows: list[dict] = []
    for k, it in enumerate(items, 1):
        tk = it["ticker"]
        row = {"ticker": tk, "name": "", "sector": "", "fit_score": None,
               "verdict": "", "risks": "", "peers": "", "n_hits": 0, "error": ""}
        try:
            data = fetch_fundamentals(tk) or {"ticker": tk}
            row["name"] = data.get("name") or ""
            row["sector"] = data.get("sector") or ""
            msg = BATCH_TEMPLATE.format(
                cot_catalog=cot_cat, note_catalog=note_cat,
                snapshot=_fundamentals_snapshot(data),
                idea=(it.get("idea") or "（无）")[:2000], top=top,
            )
            resp = client.messages.create(model=model, max_tokens=4000, system=batch_system,
                                          messages=[{"role": "user", "content": msg}])
            text = "".join(b.text for b in resp.content if b.type == "text")
            parsed = _parse_json(text)
            if not parsed:
                raise ValueError("LLM 输出无法解析为 JSON")
            try:
                row["fit_score"] = round(float(parsed.get("fit_score")), 1)
            except (TypeError, ValueError):
                row["fit_score"] = None
            row["verdict"] = str(parsed.get("verdict") or "").strip()
            row["risks"] = str(parsed.get("risks") or "").strip()
            peers = parsed.get("peers") or []
            row["peers"] = "、".join(str(x) for x in peers[:3] if x)
            hits = [h for h in (parsed.get("hits") or [])
                    if isinstance(h, dict) and isinstance(h.get("idx"), int)
                    and 0 <= h["idx"] < len(cots)][:top]
            row["n_hits"] = len(hits)
            for h in hits:
                c = cots[h["idx"]]
                hit_rows.append({
                    "ticker": tk, "cot": f"C{h['idx']}",
                    "tags": "/".join(c.get("_tags", [])), "signal": c.get("signal", ""),
                    "trigger": c.get("trigger", ""),
                    "comment": str(h.get("comment") or "").strip(),
                })
            print(f"  [{k}/{len(items)}] {tk} ✓ 契合度 {row['fit_score']} — {row['verdict'][:40]}")
        except Exception as e:
            row["error"] = str(e)
            print(f"  [{k}/{len(items)}] {tk} ✗ {e}")
        rows.append(row)

    path = None
    if save:
        out_dir = _output_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        path = out_dir / f"vet_batch_{stamp}.xlsx"
        _write_batch_xlsx(rows, hit_rows, path)
    return {"path": str(path) if path else None, "results": rows}


def _fill_sheet(ws, cols: list[tuple[str, int]], records: list[list]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    for j, (title, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        ws.column_dimensions[get_column_letter(j)].width = width
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, rec in enumerate(records, 2):
        for j, v in enumerate(rec, 1):
            ws.cell(row=i, column=j, value=v).alignment = wrap


def _write_batch_xlsx(rows: list[dict], hit_rows: list[dict], path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    _fill_sheet(ws, [("代码", 12), ("名称", 14), ("行业", 18), ("契合度/10", 10),
                     ("一句话结论", 55), ("主要风险与反逻辑", 55), ("可比同业", 20),
                     ("命中CoT数", 10), ("备注", 28)],
                [[r["ticker"], r["name"], r["sector"], r["fit_score"], r["verdict"],
                  r["risks"], r["peers"], r["n_hits"], r["error"]] for r in rows])
    ws2 = wb.create_sheet("命中明细")
    _fill_sheet(ws2, [("代码", 12), ("CoT", 8), ("主题", 20), ("信号", 7),
                      ("触发逻辑", 60), ("点评", 45)],
                [[h["ticker"], h["cot"], h["tags"], h["signal"], h["trigger"], h["comment"]]
                 for h in hit_rows])
    wb.save(str(path))
