"""Excel 抽文 — openpyxl. 每个 sheet 转为 markdown 表."""

from pathlib import Path


def load_xlsx(path: Path) -> tuple[str, int]:
    """返回 (纯文本, sheet 数)。每张表头 + 前 200 行。"""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        parts.append(f"## Sheet: {name}")
        # 截断过长 sheet
        sample = rows[:200]
        for r in sample:
            cells = []
            for v in r:
                if v is None:
                    cells.append("")
                elif isinstance(v, float):
                    # 简化浮点数显示
                    if v.is_integer():
                        cells.append(str(int(v)))
                    else:
                        cells.append(f"{v:.4g}")
                else:
                    cells.append(str(v).strip().replace("\n", " "))
            parts.append("| " + " | ".join(cells) + " |")
        if len(rows) > 200:
            parts.append(f"\n_(共 {len(rows)} 行，已截取前 200 行)_")

    wb.close()
    return "\n".join(parts), len(wb.sheetnames)
